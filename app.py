from __future__ import annotations

import io
import base64
import json
import os
import re
import textwrap
import unicodedata
import urllib.parse
import urllib.request
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, asdict
from html import escape, unescape

from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import load_workbook
from reportlab.lib.colors import HexColor
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 15 * 1024 * 1024

SAFARI_API = "https://api.distribuidorasafari.com.br/api/produto/pesquisar"
UA = "Mozilla/5.0 (compatible; SafariFotos/1.0)"

MANUFACTURERS = {
    "NATICON": "naticonindustria.com.br",
    "ASTRA": "astra-sa.com",
    "MAX": "maxeb.com.br",
    "SFORPLAST": "sforplast.com.br",
}


@dataclass
class Result:
    row: int
    code: str
    description: str
    supplier: str
    image_url: str = ""
    product_url: str = ""
    source: str = ""
    status: str = "Não encontrado"
    confidence: int = 0
    normal_price: float | None = None
    promo_price: float | None = None


def fetch(url: str, *, timeout: int = 20) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": UA, "Accept": "application/json,text/html,*/*"})
    with urllib.request.urlopen(req, timeout=timeout) as response:
        return response.read()


def norm(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode()
    return re.sub(r"[^A-Z0-9]+", " ", text.upper()).strip()


def similarity(a: str, b: str) -> int:
    sa, sb = set(norm(a).split()), set(norm(b).split())
    if not sa or not sb:
        return 0
    return round(100 * len(sa & sb) / len(sa | sb))


def safari_search(code: str, description: str) -> Result | None:
    for keyword in (code, description):
        query = urllib.parse.urlencode({"pagina": 1, "resultados_por_pagina": 12, "keyword": keyword})
        try:
            payload = json.loads(fetch(f"{SAFARI_API}?{query}"))
        except Exception:
            continue
        candidates = payload.get("resultados") or []
        if not candidates:
            continue
        exact = [p for p in candidates if str(p.get("codigo", "")).split(".")[0] == code]
        product = exact[0] if exact else max(candidates, key=lambda p: similarity(description, p.get("nome", "")))
        image = product.get("imagem") or ""
        if image and "sem_imagem" not in image.lower():
            score = 100 if exact else similarity(description, product.get("nome", ""))
            return Result(0, code, description, "", image, f"https://distribuidorasafari.com.br/produto/{product['id']}", "Safari", "Encontrado" if score >= 80 else "Revisar", score)
    return None


def manufacturer_domain(supplier: str, description: str) -> str:
    haystack = norm(f"{supplier} {description}")
    for name, domain in MANUFACTURERS.items():
        if name in haystack:
            return domain
    return ""


def manufacturer_search(description: str, supplier: str) -> Result | None:
    domain = manufacturer_domain(supplier, description)
    query = f'"{description}"' + (f" site:{domain}" if domain else " fabricante")
    url = "https://www.bing.com/images/search?" + urllib.parse.urlencode({"q": query, "form": "HDRSC2"})
    try:
        html = fetch(url).decode("utf-8", "ignore")
        for raw in re.findall(r'class="iusc"[^>]+m="([^"]+)"', html):
            meta = json.loads(unescape(raw))
            image, page = meta.get("murl", ""), meta.get("purl", "")
            if image and (not domain or domain in page):
                return Result(0, "", description, supplier, image, page, "Fabricante", "Revisar", 65)
    except Exception:
        pass
    return None


def parse_workbook(data: bytes) -> list[Result]:
    workbook = load_workbook(io.BytesIO(data), data_only=False)
    values_workbook = load_workbook(io.BytesIO(data), data_only=True)
    sheet = workbook.active
    values_sheet = values_workbook.active
    headers = {norm(cell.value): cell.column for cell in sheet[1] if cell.value}
    code_col = headers.get("PRODUTO")
    desc_col = headers.get("DESCRICAO DO PRODUTO")
    normal_col = headers.get("VENDA")
    promo_col = headers.get("PROMO")
    if not code_col or not desc_col:
        raise ValueError("A planilha precisa das colunas Produto e Descrição do Produto.")
    supplier = ""
    rows = []
    for index in range(2, sheet.max_row + 1):
        code = sheet.cell(index, code_col).value
        description = sheet.cell(index, desc_col).value
        first = sheet.cell(index, 1).value
        if first and code is None and description is None:
            supplier = str(first)
            continue
        if code is not None and description:
            normal_price = values_sheet.cell(index, normal_col).value if normal_col else None
            promo_price = values_sheet.cell(index, promo_col).value if promo_col else None
            rows.append(Result(index, str(code).strip(), str(description).strip(), supplier,
                               normal_price=float(normal_price) if isinstance(normal_price, (int, float)) else None,
                               promo_price=float(promo_price) if isinstance(promo_price, (int, float)) else None))
    return rows


@app.get("/")
def index():
    return render_template("index.html")


@app.post("/api/import")
def import_sheet():
    file = request.files.get("file")
    if not file:
        return jsonify(error="Selecione uma planilha."), 400
    try:
        rows = parse_workbook(file.read())
        return jsonify(items=[asdict(row) for row in rows], total=len(rows))
    except Exception as exc:
        return jsonify(error=str(exc)), 400


@app.post("/api/search")
def search():
    item = request.get_json(force=True)
    base = Result(int(item.get("row", 0)), str(item["code"]), item["description"], item.get("supplier", ""),
                  normal_price=item.get("normal_price"), promo_price=item.get("promo_price"))
    found = safari_search(base.code, base.description)
    if not found:
        found = manufacturer_search(base.description, base.supplier)
    if found:
        found.row, found.code, found.supplier = base.row, base.code, base.supplier
        found.normal_price, found.promo_price = base.normal_price, base.promo_price
        return jsonify(asdict(found))
    return jsonify(asdict(base))


@app.post("/api/export")
def export():
    body = request.get_json(force=True)
    items = body.get("items", [])
    original = bytes.fromhex(body["workbook_hex"])
    workbook = load_workbook(io.BytesIO(original))
    sheet = workbook.active
    start = sheet.max_column + 1
    for offset, title in enumerate(("URL da Imagem", "Fonte da Imagem", "Página do Produto", "Status da Foto")):
        sheet.cell(1, start + offset, title)
    for item in items:
        row = int(item["row"])
        sheet.cell(row, start, item.get("image_url", ""))
        sheet.cell(row, start + 1, item.get("source", ""))
        sheet.cell(row, start + 2, item.get("product_url", ""))
        sheet.cell(row, start + 3, item.get("status", ""))
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="planilha_com_fotos.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.post("/api/zip")
def images_zip():
    items = request.get_json(force=True).get("items", [])
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        for item in items:
            url, code = item.get("image_url"), item.get("code")
            if not url:
                continue
            try:
                content = fetch(url)
                ext = os.path.splitext(urllib.parse.urlparse(url).path)[1].lower()
                if ext not in {".jpg", ".jpeg", ".png", ".webp"}:
                    ext = ".jpg"
                archive.writestr(f"{code}{ext}", content)
            except Exception:
                continue
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="fotos_produtos.zip", mimetype="application/zip")


def money(value: object) -> tuple[str, str]:
    try:
        integer, cents = f"{float(value):.2f}".split(".")
        return integer, cents
    except (TypeError, ValueError):
        return "0", "00"


def section_name(supplier: str) -> str:
    cleaned = re.sub(r"^\d+\s+", "", supplier or "Ofertas Safari").strip()
    aliases = {"NATICON": "Elétrica Naticon", "ASTRA": "Sanitários Astra", "MAX METALURGICA": "Ferramentas de Jardim e Obra Max", "ELITE": "Iluminação Elite", "SFORPLAST": "Pregos Sforplast", "FARROUPILHA": "Fixação Farroupilha"}
    normalized = norm(cleaned)
    for key, label in aliases.items():
        if key in normalized:
            return label
    if "MAC FER" in normalized:
        return "Ferramentas Thompson"
    return cleaned.title()


def embed_images(items: list[dict]) -> list[dict]:
    prepared = [dict(item) for item in items]
    urls = {item.get("image_url") for item in prepared if item.get("image_url") and not item.get("image_url", "").startswith("data:")}

    def download(url: str) -> tuple[str, str]:
        try:
            content = fetch(url, timeout=8)
            path = urllib.parse.urlparse(url).path.lower()
            mime = "image/png" if path.endswith(".png") else "image/webp" if path.endswith(".webp") else "image/jpeg"
            return url, f"data:{mime};base64,{base64.b64encode(content).decode('ascii')}"
        except Exception:
            return url, url

    embedded = {}
    with ThreadPoolExecutor(max_workers=12) as pool:
        futures = [pool.submit(download, url) for url in urls]
        for future in as_completed(futures):
            url, data = future.result()
            embedded[url] = data
    for item in prepared:
        url = item.get("image_url")
        if url in embedded:
            item["image_url"] = embedded[url]
    return prepared


def paginate_tabloid(items: list[dict], columns: int) -> list[list[dict]]:
    pages: list[list[dict]] = []
    current: list[dict] = []

    def estimated_height(page_items: list[dict]) -> float:
        groups: list[tuple[str, int]] = []
        for item in page_items:
            section = section_name(item.get("supplier", ""))
            if groups and groups[-1][0] == section:
                groups[-1] = (section, groups[-1][1] + 1)
            else:
                groups.append((section, 1))
        card_rows = sum((count + columns - 1) // columns for _, count in groups)
        return len(groups) * 9.3 + card_rows * 45.5

    for item in items:
        candidate = current + [item]
        if current and estimated_height(candidate) > 238:
            pages.append(current)
            current = [item]
        else:
            current = candidate
    if current or not pages:
        pages.append(current)
    return pages


def build_tabloid(items: list[dict], start_date: str, end_date: str, density: str) -> str:
    columns = {"compacto": 5, "equilibrado": 4, "destaque": 3}.get(density, 4)
    pages = paginate_tabloid(items, columns)
    page_html = []
    for page_number, page_items in enumerate(pages, 1):
        cards, current_section = [], None
        for item in page_items:
            section = section_name(item.get("supplier", ""))
            if section != current_section:
                cards.append(f'<div class="section-title"><span>{escape(section)}</span></div>')
                current_section = section
            normal_i, normal_c = money(item.get("normal_price"))
            promo_i, promo_c = money(item.get("promo_price"))
            image = escape(item.get("image_url") or "")
            photo = f'<img class="product-photo" src="{image}" alt="Produto">' if image else '<div class="no-photo">SEM FOTO</div>'
            cards.append(f'<article class="card">{photo}<div class="card-body"><div class="cod">Cód. {escape(str(item.get("code", "")))}</div><div class="desc">{escape(item.get("description", ""))}</div></div><div class="price-row"><span class="de">De R$ {normal_i},{normal_c}</span><span class="por">R$ <b>{promo_i}</b><sup>,{promo_c}</sup></span></div></article>')
        page_html.append(f'<section class="a4-page" style="--cols:{columns}"><header><div class="hero-title">SUPER<small>OFERTAS</small></div><div class="brand"><b>Distribuidora <em>Safari</em></b><span>VÁLIDA DE {escape(start_date)} ATÉ {escape(end_date)}</span></div></header><main class="catalog">{"".join(cards)}</main><footer><span>Após a validade, os preços voltarão ao normal. Imagens meramente ilustrativas.</span><b>(11) 2911-9888 · @distribuidorasafari</b><i>{page_number}/{len(pages)}</i></footer></section>')
    styles = '''@page{size:A4 portrait;margin:0}*{box-sizing:border-box}body{margin:0;background:#dfe8e1;font-family:Arial,sans-serif;color:#183024}.a4-page{width:210mm;height:297mm;margin:10mm auto;background:#fff;display:flex;flex-direction:column;overflow:hidden;page-break-after:always}header{height:29mm;background:linear-gradient(135deg,#0d5c2e,#1a8a44);padding:6mm 8mm;display:flex;justify-content:space-between;align-items:center;color:#fff}.hero-title{font-size:28pt;font-weight:1000;line-height:.75;color:#f4e4a8}.hero-title small{display:block;font-size:11pt;letter-spacing:4px;margin-top:4mm}.brand{text-align:right;display:flex;flex-direction:column;gap:3mm}.brand b{font-size:16pt}.brand em{color:#d4af37;font-style:normal}.brand span{background:#d4af37;color:#0d5c2e;padding:2mm 4mm;border-radius:99px;font-weight:800;font-size:8pt}.catalog{padding:4mm 6mm;display:grid;grid-template-columns:repeat(var(--cols),1fr);grid-auto-flow:row;gap:2.3mm;align-content:start;flex:1;overflow:hidden}.section-title{grid-column:1/-1;border-bottom:1.2mm solid #0d5c2e;height:7mm;display:flex;align-items:end}.section-title span{background:#0d5c2e;color:#fff;padding:1.2mm 4mm;font-size:7.5pt;font-weight:900;text-transform:uppercase;border-radius:2mm 2mm 0 0}.card{height:43.2mm;border:.3mm solid #d9e3dc;border-radius:2mm;padding:2mm;display:flex;flex-direction:column;overflow:hidden;break-inside:avoid;background:#fff}.product-photo,.no-photo{width:100%;height:17mm;min-height:17mm;object-fit:contain;display:block}.no-photo{display:grid;place-items:center;background:#f1f3f1;color:#9aa39c;font-size:7pt;font-weight:bold}.card-body{height:12mm;min-height:12mm;overflow:hidden}.cod{font-size:6pt;color:#6b7c72;font-weight:bold}.desc{font-size:6.6pt;font-weight:800;line-height:1.15;margin-top:1mm;display:-webkit-box;-webkit-line-clamp:3;-webkit-box-orient:vertical;overflow:hidden}.price-row{height:10mm;min-height:10mm;border-top:.3mm dashed #d9e3dc;margin-top:1mm;padding-top:1mm}.de{display:block;font-size:6pt;color:#6b7c72;text-decoration:line-through}.por{display:block;color:#d81f2b;font-size:12pt;font-weight:1000;line-height:1}.por sup{font-size:7pt}footer{height:12mm;border-top:1mm solid #0d5c2e;margin:0 6mm;padding:2.5mm 0;display:flex;align-items:center;gap:4mm;font-size:6pt;color:#6b7c72}footer b{color:#183024;margin-left:auto}footer i{font-style:normal;font-weight:bold}@media print{body{background:#fff}.a4-page{margin:0}}@media screen{.a4-page{box-shadow:0 8px 32px #0002}}'''
    return f'<!doctype html><html lang="pt-BR"><head><meta charset="utf-8"><title>Tabloide Safari</title><style>{styles}</style></head><body>{"".join(page_html)}</body></html>'


def svg_text_lines(text: str, width: int = 28, limit: int = 3) -> list[str]:
    lines = textwrap.wrap(str(text), width=width, break_long_words=False, break_on_hyphens=False)
    if len(lines) > limit:
        lines = lines[:limit]
        lines[-1] = lines[-1].rstrip(" .") + "…"
    return lines or [""]


def build_vector_pages(items: list[dict], start_date: str, end_date: str, density: str) -> list[str]:
    columns = {"compacto": 5, "equilibrado": 4, "destaque": 3}.get(density, 4)
    pages = paginate_tabloid(items, columns)
    width, height, margin, gap = 794, 1123, 24, 8
    card_width = (width - margin * 2 - gap * (columns - 1)) / columns
    card_height, section_height = 163, 29
    output = []
    for page_number, page_items in enumerate(pages, 1):
        elements = [f'''<rect width="794" height="1123" fill="#ffffff"/><defs><linearGradient id="hero" x1="0" y1="0" x2="1" y2="1"><stop stop-color="#0d5c2e"/><stop offset="1" stop-color="#1a8a44"/></linearGradient><clipPath id="photo"><rect width="{card_width-14:.1f}" height="62" rx="4"/></clipPath></defs><rect width="794" height="110" fill="url(#hero)"/><text x="30" y="58" font-family="Arial" font-size="43" font-weight="900" fill="#f4e4a8">SUPER</text><text x="32" y="87" font-family="Arial" font-size="18" font-weight="800" letter-spacing="5" fill="#ffffff">OFERTAS</text><text x="764" y="48" text-anchor="end" font-family="Arial" font-size="24" font-weight="800" fill="#ffffff">Distribuidora <tspan fill="#d4af37">Safari</tspan></text><rect x="485" y="64" width="279" height="27" rx="14" fill="#d4af37"/><text x="624" y="82" text-anchor="middle" font-family="Arial" font-size="11" font-weight="800" fill="#0d5c2e">VÁLIDA DE {escape(start_date)} ATÉ {escape(end_date)}</text>''']
        y, index, current_section = 126, 0, None
        while index < len(page_items):
            section = section_name(page_items[index].get("supplier", ""))
            if section != current_section:
                elements.extend([f'<rect x="{margin}" y="{y}" width="{width-margin*2}" height="5" fill="#0d5c2e"/>', f'<rect x="{margin}" y="{y-21}" width="260" height="26" rx="5" fill="#0d5c2e"/>', f'<text x="{margin+12}" y="{y-4}" font-family="Arial" font-size="11" font-weight="800" fill="#ffffff">{escape(section.upper())}</text>'])
                y += section_height
                current_section = section
            group = []
            while index < len(page_items) and section_name(page_items[index].get("supplier", "")) == section and len(group) < columns:
                group.append(page_items[index]); index += 1
            for col, item in enumerate(group):
                x = margin + col * (card_width + gap)
                code = escape(str(item.get("code", "")))
                normal_i, normal_c = money(item.get("normal_price")); promo_i, promo_c = money(item.get("promo_price"))
                elements.append(f'<rect x="{x:.1f}" y="{y}" width="{card_width:.1f}" height="{card_height}" rx="7" fill="#ffffff" stroke="#d9e3dc"/>')
                image = item.get("image_url") or ""
                if image:
                    elements.append(f'<image href="{escape(image)}" x="{x+7:.1f}" y="{y+7}" width="{card_width-14:.1f}" height="62" preserveAspectRatio="xMidYMid meet"/>')
                else:
                    elements.extend([f'<rect x="{x+7:.1f}" y="{y+7}" width="{card_width-14:.1f}" height="62" fill="#f1f3f1"/>', f'<text x="{x+card_width/2:.1f}" y="{y+42}" text-anchor="middle" font-family="Arial" font-size="9" fill="#9aa39c">SEM FOTO</text>'])
                elements.append(f'<text x="{x+8:.1f}" y="{y+81}" font-family="Arial" font-size="8" font-weight="700" fill="#6b7c72">Cód. {code}</text>')
                for line_no, line in enumerate(svg_text_lines(item.get("description", ""), 25 if columns >= 4 else 34)):
                    elements.append(f'<text x="{x+8:.1f}" y="{y+96+line_no*10}" font-family="Arial" font-size="8.5" font-weight="700" fill="#183024">{escape(line)}</text>')
                elements.extend([f'<line x1="{x+8:.1f}" y1="{y+128}" x2="{x+card_width-8:.1f}" y2="{y+128}" stroke="#d9e3dc" stroke-dasharray="3 3"/>', f'<text x="{x+8:.1f}" y="{y+140}" font-family="Arial" font-size="8" fill="#6b7c72" text-decoration="line-through">De R$ {normal_i},{normal_c}</text>', f'<text x="{x+8:.1f}" y="{y+157}" font-family="Arial" font-size="16" font-weight="900" fill="#d81f2b">R$ {promo_i}<tspan font-size="10">,{promo_c}</tspan></text>'])
            y += card_height + gap
        elements.extend([f'<line x1="24" y1="1076" x2="770" y2="1076" stroke="#0d5c2e" stroke-width="4"/>', '<text x="24" y="1095" font-family="Arial" font-size="8" fill="#6b7c72">Após a validade, os preços voltarão ao normal. Imagens meramente ilustrativas.</text>', '<text x="660" y="1095" text-anchor="end" font-family="Arial" font-size="8" font-weight="700" fill="#183024">(11) 2911-9888 · @distribuidorasafari</text>', f'<text x="770" y="1095" text-anchor="end" font-family="Arial" font-size="8" font-weight="700" fill="#183024">{page_number}/{len(pages)}</text>'])
        output.append(f'''<svg xmlns="http://www.w3.org/2000/svg" xmlns:xlink="http://www.w3.org/1999/xlink" width="210mm" height="297mm" viewBox="0 0 794 1123">{"".join(elements)}</svg>''')
    return output


def image_reader(source: str) -> ImageReader | None:
    if not source:
        return None
    try:
        if source.startswith("data:"):
            raw = base64.b64decode(source.split(",", 1)[1])
        else:
            raw = fetch(source, timeout=8)
        return ImageReader(io.BytesIO(raw))
    except Exception:
        return None


def build_illustrator_pages(items: list[dict], start_date: str, end_date: str, density: str) -> list[bytes]:
    columns = {"compacto": 5, "equilibrado": 4, "destaque": 3}.get(density, 4)
    pages = paginate_tabloid(items, columns)
    page_w, page_h = A4
    margin, gap, header_h, footer_h = 18, 6, 82, 34
    card_w = (page_w - margin * 2 - gap * (columns - 1)) / columns
    card_h, section_h = 122, 22
    results = []

    def top_y(top: float, height: float = 0) -> float:
        return page_h - top - height

    for page_number, page_items in enumerate(pages, 1):
        output = io.BytesIO()
        pdf = canvas.Canvas(output, pagesize=A4, pageCompression=1)
        pdf.setTitle("Tabloide Safari Editável")
        pdf.setAuthor("Distribuidora Safari")

        pdf.setFillColor(HexColor("#0d5c2e")); pdf.rect(0, top_y(0, header_h), page_w, header_h, fill=1, stroke=0)
        pdf.setFillColor(HexColor("#f4e4a8")); pdf.setFont("Helvetica-Bold", 31); pdf.drawString(24, top_y(42), "SUPER")
        pdf.setFillColor(HexColor("#ffffff")); pdf.setFont("Helvetica-Bold", 13); pdf.drawString(25, top_y(65), "O F E R T A S")
        pdf.setFont("Helvetica-Bold", 18); pdf.drawRightString(page_w-24, top_y(35), "Distribuidora Safari")
        pdf.setFillColor(HexColor("#d4af37")); pdf.roundRect(page_w-232, top_y(67, 20), 208, 20, 10, fill=1, stroke=0)
        pdf.setFillColor(HexColor("#0d5c2e")); pdf.setFont("Helvetica-Bold", 8); pdf.drawCentredString(page_w-128, top_y(80), f"VÁLIDA DE {start_date} ATÉ {end_date}")

        y, index, current_section = header_h + 13, 0, None
        while index < len(page_items):
            section = section_name(page_items[index].get("supplier", ""))
            if section != current_section:
                pdf.setFillColor(HexColor("#0d5c2e")); pdf.roundRect(margin, top_y(y, 19), 210, 19, 4, fill=1, stroke=0)
                pdf.setFont("Helvetica-Bold", 8); pdf.setFillColor(HexColor("#ffffff")); pdf.drawString(margin+9, top_y(y+13), section.upper())
                y += section_h; current_section = section
            group = []
            while index < len(page_items) and section_name(page_items[index].get("supplier", "")) == section and len(group) < columns:
                group.append(page_items[index]); index += 1
            for col, item in enumerate(group):
                x = margin + col * (card_w + gap); bottom = top_y(y, card_h)
                pdf.setFillColor(HexColor("#ffffff")); pdf.setStrokeColor(HexColor("#d9e3dc")); pdf.roundRect(x, bottom, card_w, card_h, 5, fill=1, stroke=1)
                reader = image_reader(item.get("image_url") or "")
                photo_x, photo_y, photo_w, photo_h = x+6, top_y(y+5, 48), card_w-12, 48
                if reader:
                    iw, ih = reader.getSize(); scale = min(photo_w/iw, photo_h/ih)
                    dw, dh = iw*scale, ih*scale
                    pdf.drawImage(reader, photo_x+(photo_w-dw)/2, photo_y+(photo_h-dh)/2, dw, dh, preserveAspectRatio=True, mask="auto")
                else:
                    pdf.setFillColor(HexColor("#f1f3f1")); pdf.rect(photo_x, photo_y, photo_w, photo_h, fill=1, stroke=0)
                    pdf.setFillColor(HexColor("#9aa39c")); pdf.setFont("Helvetica-Bold", 7); pdf.drawCentredString(x+card_w/2, photo_y+22, "SEM FOTO")
                pdf.setFillColor(HexColor("#6b7c72")); pdf.setFont("Helvetica-Bold", 6); pdf.drawString(x+6, top_y(y+62), f"Cód. {item.get('code','')}")
                pdf.setFillColor(HexColor("#183024")); pdf.setFont("Helvetica-Bold", 6.4)
                for line_no, line in enumerate(svg_text_lines(item.get("description", ""), 24 if columns >= 4 else 32, 3)):
                    pdf.drawString(x+6, top_y(y+74+line_no*8), line)
                pdf.setStrokeColor(HexColor("#d9e3dc")); pdf.setDash(2, 2); pdf.line(x+6, top_y(y+98), x+card_w-6, top_y(y+98)); pdf.setDash()
                normal_i, normal_c = money(item.get("normal_price")); promo_i, promo_c = money(item.get("promo_price"))
                pdf.setFillColor(HexColor("#6b7c72")); pdf.setFont("Helvetica", 6); pdf.drawString(x+6, top_y(y+108), f"De R$ {normal_i},{normal_c}")
                pdf.setFillColor(HexColor("#d81f2b")); pdf.setFont("Helvetica-Bold", 13); pdf.drawString(x+6, top_y(y+120), f"R$ {promo_i},{promo_c}")
            y += card_h + gap

        footer_y = footer_h
        pdf.setStrokeColor(HexColor("#0d5c2e")); pdf.setLineWidth(3); pdf.line(margin, footer_y, page_w-margin, footer_y)
        pdf.setFillColor(HexColor("#6b7c72")); pdf.setFont("Helvetica", 5.8); pdf.drawString(margin, 17, "Após a validade, os preços voltarão ao normal. Imagens meramente ilustrativas.")
        pdf.setFillColor(HexColor("#183024")); pdf.setFont("Helvetica-Bold", 6); pdf.drawRightString(page_w-42, 17, "(11) 2911-9888 · @distribuidorasafari")
        pdf.drawRightString(page_w-margin, 17, f"{page_number}/{len(pages)}")
        pdf.showPage(); pdf.save(); results.append(output.getvalue())
    return results


@app.post("/api/tabloid")
def tabloid():
    body = request.get_json(force=True)
    items = embed_images(body.get("items", []))
    html = build_tabloid(items, body.get("start_date", ""), body.get("end_date", ""), body.get("density", "equilibrado"))
    return send_file(io.BytesIO(html.encode("utf-8")), as_attachment=True, download_name="tabloide_safari.html", mimetype="text/html; charset=utf-8")


@app.post("/api/vector")
def vector():
    body = request.get_json(force=True)
    items = embed_images(body.get("items", []))
    pages = build_illustrator_pages(items, body.get("start_date", ""), body.get("end_date", ""), body.get("density", "equilibrado"))
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        for index, ai in enumerate(pages, 1):
            archive.writestr(f"tabloide_safari_pagina_{index:02d}.ai", ai)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="tabloide_safari_editavel.zip", mimetype="application/zip")


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", "5000")), debug=True)
